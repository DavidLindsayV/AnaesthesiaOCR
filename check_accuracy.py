import re
import sys
from nltk.metrics.distance import edit_distance
import csv
from jiwer import cer
import matplotlib.pyplot as plt
import os
import openpyxl

from monitor_values import Field_Ranges, HospitalMonitor, OldMonitor


def get_extracted_data(extracted_data_path, fields):
    extracted_data = []
    with open(extracted_data_path, "r") as file:
        csvreader = csv.reader(file)
        firstrow = True
        columns = []
        for row in csvreader:
            if firstrow:
                firstrow = False
                columns = row
                continue
            processed_row = [None for x in range(len(fields))]
            for i, field in enumerate(columns):
                if field in fields:
                    processed_row[fields.index(field)] = str(row[i])
            extracted_data.append(processed_row)

    print("Extracted data loaded")
    return extracted_data


def get_expected_data_row(expected_data_sheet, fields, row_num):
    row_num = int(row_num)
    dataframe = openpyxl.load_workbook(os.path.join("images", "monitor_data.xlsx"))
    dataframe1 = dataframe[expected_data_sheet]
    firstrow = True
    end_of_data_reached = False
    columns = []
    expected_data = []
    for row in range(0, dataframe1.max_row):
        if firstrow:
            firstrow = False
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                columns.append(col[row].value)
            continue
        if row == row_num:
            row_data = [None for x in range(len(fields))]
            row_dict = {}
            col_num = -1
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                col_num += 1
                if columns[col_num] in fields:
                    row_dict[columns[col_num]] = str(col[row].value)
            return row_dict
        if end_of_data_reached:
            break
    return None


def get_expected_data(expected_data_sheet, fields):
    dataframe = openpyxl.load_workbook(os.path.join("images", "monitor_data.xlsx"))
    dataframe1 = dataframe[expected_data_sheet]
    # expected_data = pd.read_excel(os.path.join("images","monitor_data.xlsx"), sheet_name="OldMonitor")
    firstrow = True
    end_of_data_reached = False
    columns = []
    expected_data = []
    for row in range(0, dataframe1.max_row):
        if firstrow:
            firstrow = False
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                columns.append(col[row].value)
            continue
        row_data = [None for x in range(len(fields))]
        col_num = -1
        for col in dataframe1.iter_cols(1, dataframe1.max_column):
            col_num += 1

            if columns[col_num] in fields:
                row_data[fields.index(columns[col_num])] = str(col[row].value)
        if end_of_data_reached:
            break
        expected_data.append(row_data)

    print("Expected data loaded")
    return expected_data


def is_float(string):
    string = string.replace("(", "").replace(")", "")
    try:
        float(string)
        return True
    except ValueError:
        return False

def update_eval_params_minmax(eval_params_minmax, j, index, metric):
    if metric < eval_params_minmax[j][index][0]:
        eval_params_minmax[j][index][0] = metric
    if metric > eval_params_minmax[j][index][1]:
        eval_params_minmax[j][index][1] = metric
    return eval_params_minmax

def update_eval_params(eval_params, eval_params_minmax, j, is_correct, edit_dist, cer, num_dist, num_dist_norm):
    if is_correct:
        eval_params[j][0] += 1
    eval_params[j][1] += edit_dist
    eval_params[j][2] += cer
    eval_params_minmax = update_eval_params_minmax(eval_params_minmax, j, 1, edit_dist)
    eval_params_minmax = update_eval_params_minmax(eval_params_minmax, j, 2, cer)

    if num_dist >= 0 and num_dist_norm >= 0:
        eval_params[j][3] += num_dist
        eval_params[j][4] += num_dist_norm
        eval_params_minmax = update_eval_params_minmax(eval_params_minmax, j, 3, num_dist)
        eval_params_minmax = update_eval_params_minmax(eval_params_minmax, j, 4, num_dist_norm)  

    return eval_params, eval_params_minmax      

def calculate_accuracy_metrics(extracted_data, expected_data, fields):
    num_correct = 0
    tot_edit_dist = 0
    tot_cer = 0
    tot_numerical_dist = 0
    eval_params = [
        [0 for y in range(5)] for x in range(len(fields))
    ]  # this is a 2d array. The first index is what the field is. The second array is of [count_of_correct_ocr, total_edit_distance, total_cer, tot_numeric_dist, tot_numeric_dist_normalised]
    eval_params_minmax = [[[999999999,-999999999] for y in range(5)] for x in range(len(fields))] #this is the same format as eval_params, but instead of being 1 number, it's two: the minimum and maximum value for that metric

    # Estimates for the ranges of each physiological parameter. Note, these are ESTIMATES  TODO double check these estimates??
    field_ranges = Field_Ranges.field_ranges

    for i in range(len(extracted_data)):
        for j in range(len(fields)):
            exp = expected_data[i][j]
            extr = extracted_data[i][j]

            #eval params
            is_correct = False
            edit_dist = -1
            CER = -1
            num_dist = -1
            num_dist_norm = -1

            if exp == extr:
                num_correct += 1
                is_correct = True
            else:
                print("Field: " + fields[j] + " Exp: " + exp + " Extr: " + extr)
            tot_edit_dist += edit_distance(exp, extr)
            tot_cer += cer(exp, extr)
            edit_dist = edit_distance(exp, extr)
            CER = cer(exp, extr)

            if is_float(extr) and is_float(exp):
                exp = re.sub(r"[^0-9.]", "", exp)
                extr = re.sub(r"[^0-9.]", "", extr)
                tot_numerical_dist += abs(float(exp) - float(extr))
                num_dist = abs(float(exp) - float(extr))
                field_range = field_ranges[fields[j]][0] - field_ranges[fields[j]][1]
                num_dist_norm = abs(float(exp) - float(extr)) / field_range
            
            eval_params, eval_params_minmax = update_eval_params(eval_params, eval_params_minmax, j, is_correct, edit_dist, CER, num_dist, num_dist_norm)

        print("Processed " + str(i + 1) + " out of " + str(len(expected_data)))

    tot_param_count = len(extracted_data) * len(fields)
    avg_accuracy = "{:.1f}".format((num_correct / tot_param_count) * 100)
    avg_edit_distance = "{:.3f}".format((tot_edit_dist / tot_param_count))
    avg_cer = "{:.3f}".format((tot_cer / tot_param_count) * 100)
    avg_numerical_distance = "{:.3f}".format((tot_numerical_dist / tot_param_count))
    image_count = len(extracted_data)

    return (
        avg_accuracy,
        avg_edit_distance,
        avg_cer,
        avg_numerical_distance,
        image_count,
        eval_params,
        eval_params_minmax
    )


def print_accuracy_metrics(
    avg_accuracy,
    avg_edit_distance,
    avg_cer,
    avg_numerical_distance,
    eval_params,
    eval_params_minmax,
    image_count,
    fields,
):
    print()
    print("ACCURACY: " + avg_accuracy + "%")
    print("AVG EDIT DISTANCE: " + avg_edit_distance + " edits")
    print("AVG CHARACTER ERROR RATE: " + avg_cer + "%")
    print("AVG NUMERICAL DISTANCE: " + avg_numerical_distance)
    print()
    for i in range(len(fields)):

        eval_params[i][0] = "{:.1f}".format((eval_params[i][0] / image_count) * 100)
        eval_params[i][1] = "{:.3f}".format((eval_params[i][1] / image_count))
        eval_params_minmax[i][1][0] = "{:}".format(eval_params_minmax[i][1][0])
        eval_params_minmax[i][1][1] = "{:}".format(eval_params_minmax[i][1][1])
        eval_params[i][2] = "{:.3f}".format((eval_params[i][2] / image_count) * 100)
        eval_params_minmax[i][2][0] = "{:.0f}".format(eval_params_minmax[i][2][0]*100)
        eval_params_minmax[i][2][1] = "{:.0f}".format(eval_params_minmax[i][2][1]*100)
        eval_params[i][3] = "{:.1f}".format((eval_params[i][3] / image_count))
        eval_params_minmax[i][3][0] = "{:.0f}".format(eval_params_minmax[i][3][0])
        eval_params_minmax[i][3][1] = "{:.0f}".format(eval_params_minmax[i][3][1])
        eval_params[i][4] = "{:.1f}".format(((eval_params[i][4]) / image_count) * 100)
        eval_params_minmax[i][4][0] = "{:.0f}".format(eval_params_minmax[i][4][0] * 100)
        eval_params_minmax[i][4][1] = "{:.0f}".format(eval_params_minmax[i][4][1] * 100)
        print(
            "Field "
            + str(fields[i])
            + " Accuracy = "
            + eval_params[i][0]
            + "% | Avg Edit Distance: "
            + eval_params[i][1]
            + " edits (min " + eval_params_minmax[i][1][0] + ", max " + eval_params_minmax[i][1][1] + ") | Avg CER: "
            + eval_params[i][2]
            + "% (min " + eval_params_minmax[i][2][0] + ", max " + eval_params_minmax[i][2][1] + ") | Avg Numeric Distance = "
            + eval_params[i][3]
            + " (min " + eval_params_minmax[i][3][0] + ", max " + eval_params_minmax[i][3][1] + ") | Normalised Avg Numeric Distance = "
            + eval_params[i][4]
            + "% (min " + eval_params_minmax[i][4][0] + ", max " + eval_params_minmax[i][4][1] + ")"
        )


def create_accuracy_pyplot(
    avg_accuracy,
    avg_edit_distance,
    avg_cer,
    avg_numerical_distance,
    eval_params,
    eval_params_minmax,
    fields,
):
    for i in range(len(fields)):
        #Update the eval_params to the best format for the pyplot
        eval_params[i][1] = eval_params[i][1] + " (min " + eval_params_minmax[i][1][0] + ", max " + eval_params_minmax[i][1][1] + ")"
        eval_params[i][2] = eval_params[i][2] + " (min " + eval_params_minmax[i][2][0] + ", max " + eval_params_minmax[i][2][1] + ")"
        eval_params[i][3] = eval_params[i][3] + " (min " + eval_params_minmax[i][3][0] + ", max " + eval_params_minmax[i][3][1] + ")"
        eval_params[i][4] = eval_params[i][4] + " (min " + eval_params_minmax[i][4][0] + ", max " + eval_params_minmax[i][4][1] + ")"

    fig, ax = plt.subplots()
    ax.xaxis.set_visible(False)
    ax.yaxis.set_visible(False)
    ax.set_frame_on(False)
    columns = [
        "Accuracy (%)",
        "Avg Edit Distance (edits)",
        "Avg Character Error Rate (%)",
        "Avg Numeric Distance (unitless)",
        "Norm Avg Numeric Distance (%)",
    ]
    fields.append("Average")
    rows = fields
    eval_params.append(
        [avg_accuracy, avg_edit_distance, avg_cer, avg_numerical_distance, "NA"]
    )
    table_data = eval_params
    table = ax.table(
        cellText=table_data,
        rowLabels=rows,
        colLabels=columns,
        cellLoc="center",
        loc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(10)  # Set the desired font size
    table.scale(1, 1)  # Scale the table cells (width, height)
    for key, cell in table.get_celld().items():
        if key[0] == 0 or key[1] == -1:
            cell.set_text_props(weight="bold")
            # cell.set_text_props(fontsize='10')
            cell.set_facecolor("#D3D3D3")  # Light gray background for header
        else:
            cell.set_edgecolor("black")  # Black border for cells
            cell.set_linewidth(0.5)  # Border line width
    final_row_idx = len(rows)
    for col_idx in range(len(columns)):
        cell = table[(final_row_idx, col_idx)]
        cell.set_text_props(weight="bold")
        cell.set_linewidth(1)

    table.auto_set_column_width(col=list(range(len(columns))))  # Adjust column widths
    plt.subplots_adjust(left=0.2, top=0.8)
    plt.show()


def calculate_accuracy(extracted_data_path, expected_data_sheet):
    monitor = HospitalMonitor()
    print(expected_data_sheet)
    if expected_data_sheet == "OldMonitor":
        monitor = OldMonitor()
    fields = list(monitor.get_field_pos().keys())

    extracted_data = get_extracted_data(extracted_data_path, fields)
    expected_data = get_expected_data(expected_data_sheet, fields)
    if len(expected_data) != len(extracted_data):
        print(
            "ERROR: Length of expected data doesn't match length of extracted data. Ending program"
        )
        sys.exit()

    print(
        "Number of images in OCR output: "
        + str(len(extracted_data))
        + " \nNumber of images in testing data: "
        + str(len(expected_data))
    )

    (
        avg_accuracy,
        avg_edit_distance,
        avg_cer,
        avg_numerical_distance,
        image_count,
        eval_params,
        eval_params_minmax
    ) = calculate_accuracy_metrics(extracted_data, expected_data, fields)

    print_accuracy_metrics(
        avg_accuracy,
        avg_edit_distance,
        avg_cer,
        avg_numerical_distance,
        eval_params,
        eval_params_minmax,
        image_count,
        fields,
    )
    create_accuracy_pyplot(
        avg_accuracy,
        avg_edit_distance,
        avg_cer,
        avg_numerical_distance,
        eval_params,
        eval_params_minmax, #TODO make create_acc_pyplot and print_acc_metrics use minmax for the eval params
        fields,
    )


if __name__ == "__main__":
    if len(sys.argv) == 3:
        calculate_accuracy(sys.argv[1], sys.argv[2])
    else:
        print(
            "Need to enter 2 cmd arguments: First is the csv file of the OCR outputs, the second is what sheet in the excel file has the image data"
        )
        print(
            "Options for image data: OldMonitor, NormalHospital, RepositionedCameraHospital, DarkHospital, BrightReflectionHospital"
        )


# TODO calculate the odds of getting an entire row correct, not just one field in a row
