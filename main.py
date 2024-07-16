import os
import random
from monitor_values import HospitalMonitor, OldMonitor
from process_img import process_img
from query import extract_data
from write_to_csv import write_to_csv
from datetime import datetime
import sys


def checkAnswers(extracted_data):
    ###Tests the answers to see if the numbers extracted match image.png
    print()
    expected_data = {'ecg.hr': '76', 'co2.et': '37', 'co2.fi': '0', 'co2.rr': '16', 'p1.sys': '139', 'p1.dia': '79', 'p1.mean': '(94)', 'aa.et': '0.80', 'aa.fi': '2.1'}
    print(expected_data)
    print(extracted_data)

    for key in expected_data.keys():
        if expected_data[key] != extracted_data[key]:
            print("WRONG " + key + " Expected: " + expected_data[key] + " Actual: " + extracted_data[key])

def checkAnswersForImg(imgNum, ocrAnswers):
    import openpyxl
    dataframe = openpyxl.load_workbook("monitor_data.xlsx")
    dataframe1 = dataframe.active
    firstrow = True
    columns = []
    expected_data = {}
    for row in range(0, dataframe1.max_row):
        if firstrow:
            firstrow = False
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                columns.append(col[row].value)
            continue
        if row == imgNum:
            col_num = -1
            for col in dataframe1.iter_cols(1, dataframe1.max_column):
                col_num += 1
                if columns[col_num] in ocrAnswers.keys():
                    expected_data[columns[col_num]] = str(col[row].value)
    
    print("Image = " + str(imgNum) + "tmp.jpg")
    print(expected_data)
    print(ocrAnswers)
    for key in ocrAnswers.keys():
        if expected_data[key] != ocrAnswers[key]:
            print("WRONG " + key + " Expected: " + expected_data[key] + " Actual: " + ocrAnswers[key])

def test_with_one_image():
    imagesDict = process_img(os.path.join("misc_images", "image.png"))
    time = datetime.now()
    extracted_data = [extract_data(imagesDict)]
    print("Time taken to perform AI OCR = " + str(datetime.now() - time))
    checkAnswers(extracted_data[0])

def test_with_img(imgNum): #TODO update this to work with other image folders
    imagesDict = process_img(os.path.join("images", str(imgNum) + "tmp.jpg"))
    time = datetime.now()
    extracted_data = [extract_data(imagesDict)]
    print("Time taken to perform AI OCR = " + str(datetime.now() - time))
    checkAnswersForImg(imgNum, extracted_data[0])

def test_with_random_image():
    imgNum = random.randint(1, 182)
    test_with_img(imgNum)

def write_to_csv_all_images(img_folder):
    dir = os.path.join("images", img_folder)

    monitor = HospitalMonitor()
    if img_folder == "oldmonitor_images":
        monitor = OldMonitor()

    bbox_adjustment = True

    starttime = datetime.now()
    ocr_data = []
    num_images = len(os.listdir(dir))
    count = 1
    num_images = len(os.listdir(dir))
    for i in range(1, num_images + 1):
        print("Processing image " + str(count) + "/" + str(num_images))
        count += 1
        filename = os.path.join(dir, str(i) + "tmp.jpg")
        imagesDict = process_img(filename, monitor, bbox_adjustment)
        print(filename)
        ocr_data.append(extract_data(imagesDict))
    write_to_csv(ocr_data)
    print("Completed! Time taken = " + str(datetime.now() - starttime))

# if len(sys.argv) > 1:
#     test_with_img(int(sys.argv[1]))
# else:
#     test_with_random_image()

if len(sys.argv) >= 2:
    write_to_csv_all_images(sys.argv[1])
else:
    print("Need to know what images to make into csv. Provide one command line argument, one of the names of the image subfolders within the image folder")